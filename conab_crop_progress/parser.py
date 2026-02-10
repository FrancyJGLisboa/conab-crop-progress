"""Excel parsing for CONAB crop progress spreadsheets.

Each Excel file contains one sheet ("Progresso de safra") with multiple
crop blocks stacked vertically.  This module iterates row-by-row detecting
block boundaries and emitting flat records matching the target schema.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from .translator import normalize_activity, normalize_crop, normalize_state, state_code_to_name

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Regex patterns
# ---------------------------------------------------------------------------

# "Soja - Safra 2025/26" or "Trigo - Safra 2025"
CROP_HEADER_RE = re.compile(r"^(.+?)\s*-\s*Safra\s+(.+)$", re.IGNORECASE)

# Activity labels
ACTIVITY_RE = re.compile(r"^(Semeadura|Colheita)\s*\*?\s*$", re.IGNORECASE)

# State header variants
STATE_HEADER_RE = re.compile(
    r"^(Estado|Unidade da Federa[çc][ãa]o)$", re.IGNORECASE
)

# Total/summary row
TOTAL_ROW_RE = re.compile(r"^\d+\s*estados?$", re.IGNORECASE)

# Numeric percentage with possible annotations: "0,5% **", "100", "15% (1)"
PERCENT_RE = re.compile(r"([\d]+(?:[.,]\d+)?)\s*%?\s*(?:\**|\(.*\)|[⁽⁾\d]*)\s*$")


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

def _coerce_pct(value: Any) -> float | None:
    """Coerce a cell value to a float percentage in [0, 1].

    openpyxl returns Excel percentage-formatted cells already as decimals
    (0.56 = 56%, 1.0 = 100%), so numeric values are used as-is.

    String annotations (e.g. "0,5% **", "50%") are on a 0-100 text scale
    and must be divided by 100.

    Handles: int, float, str with commas, str with '%', None.
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        # openpyxl already returns percentages as 0-1 decimals
        return float(value)

    if isinstance(value, str):
        value = value.strip()
        if not value or value == "-":
            return None
        m = PERCENT_RE.search(value)
        if m:
            num_str = m.group(1).replace(",", ".")
            return float(num_str) / 100.0
        # Last resort: try direct parse
        try:
            return float(value.replace(",", ".")) / 100.0
        except ValueError:
            logger.warning("Could not coerce percentage value: %r", value)
            return None

    return None


def _coerce_date(value: Any) -> date | None:
    """Coerce a cell value to a date object."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


# ---------------------------------------------------------------------------
# Row reading helpers
# ---------------------------------------------------------------------------

def _row_values(ws, row_idx: int, max_col: int) -> list[Any]:
    """Read a row's cell values, resolving merged cells."""
    values: list[Any] = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        values.append(cell.value)
    return values


def _find_text_in_row(values: list[Any]) -> str | None:
    """Return the first non-None string value in a row (stripped)."""
    for v in values:
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def _find_text_at_b(values: list[Any]) -> str | None:
    """Return the string in column B (index 1), or None."""
    if len(values) > 1 and isinstance(values[1], str) and values[1].strip():
        return values[1].strip()
    return None


# ---------------------------------------------------------------------------
# Main file parser
# ---------------------------------------------------------------------------

def parse_file(filepath: Path) -> list[dict]:
    """Parse a single CONAB crop progress Excel file into flat records.

    Returns a list of dicts matching the target schema.
    """
    filepath = Path(filepath)
    logger.info("Parsing: %s", filepath.name)

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        logger.error("Failed to open %s: %s", filepath, exc)
        return []

    # Use the first sheet (usually named "Progresso de safra")
    ws = wb.active
    if ws is None:
        logger.error("No active sheet in %s", filepath)
        wb.close()
        return []

    max_col = ws.max_column or 10
    max_row = ws.max_row or 0
    if max_col < 4:
        max_col = 10  # safety: some sheets report wrong max_column

    records: list[dict] = []

    # State machine variables
    current_crop_pt: str | None = None
    current_crop_en: str | None = None
    current_season: str | None = None
    current_activity_pt: str | None = None
    current_activity_en: str | None = None
    date_columns: list[date | None] = []
    has_five_year_avg = False
    data_col_start = 2  # 0-indexed column C in values list
    expect_dates = False  # waiting for date row after header row

    for row_idx in range(1, max_row + 1):
        values = _row_values(ws, row_idx, max_col)

        # --- Try to detect crop header ---
        text = _find_text_in_row(values)
        if text:
            m = CROP_HEADER_RE.match(text)
            if m:
                crop_pt = m.group(1).strip()
                season = m.group(2).strip()
                crop_en = normalize_crop(crop_pt)
                if crop_en:
                    current_crop_pt = crop_pt
                    current_crop_en = crop_en
                    current_season = season
                    current_activity_pt = None
                    current_activity_en = None
                    date_columns = []
                    has_five_year_avg = False
                    expect_dates = False
                    logger.debug(
                        "  Crop block: %s (%s) - Safra %s",
                        crop_pt, crop_en, season,
                    )
                else:
                    logger.warning(
                        "  Unrecognized crop: %r in %s", crop_pt, filepath.name
                    )
                continue

        if current_crop_en is None:
            continue

        # --- Detect activity row ---
        text_b = _find_text_at_b(values)
        if text_b:
            m_act = ACTIVITY_RE.match(text_b)
            if m_act:
                current_activity_pt = m_act.group(1).strip()
                current_activity_en = normalize_activity(current_activity_pt)
                date_columns = []
                has_five_year_avg = False
                expect_dates = False
                logger.debug("    Activity: %s", current_activity_en)
                continue

        if current_activity_en is None:
            continue

        # --- Detect state/header row (precedes year labels and dates) ---
        if text_b and STATE_HEADER_RE.match(text_b):
            # Check if "Média 5 anos" appears in this row
            row_text = " ".join(str(v) for v in values if v)
            has_five_year_avg = "dia" in row_text.lower() and "ano" in row_text.lower()
            expect_dates = True
            continue

        # --- Detect year label row (just years like 2025, 2026) ---
        # These are integer/float years; skip but note them
        if expect_dates:
            numeric_count = sum(
                1 for v in values[2:] if isinstance(v, (int, float)) and 2000 <= v <= 2100
            )
            if numeric_count >= 2:
                # This is the year-label row; dates come next
                continue

        # --- Detect date row ---
        if expect_dates:
            dates_found = [_coerce_date(v) for v in values[2:]]
            actual_dates = [d for d in dates_found if d is not None]
            if actual_dates:
                date_columns = dates_found
                expect_dates = False
                logger.debug(
                    "    Dates: %s",
                    [str(d) for d in actual_dates],
                )
                continue

        # --- Detect data rows (state rows or total) ---
        if not date_columns:
            continue

        state_name = text_b
        if state_name is None:
            continue

        # Check for total row
        is_total = bool(TOTAL_ROW_RE.match(state_name))

        if not is_total:
            state_code = normalize_state(state_name)
            if state_code is None:
                # Might be a spurious row or unrecognized label
                if state_name.lower() not in ("", "fonte:", "nota:", "obs:"):
                    logger.debug(
                        "    Skipping unrecognized row: %r", state_name
                    )
                continue
        else:
            state_code = None

        # Extract numeric values from data columns (starting at col C = index 2)
        data_values = values[data_col_start:]

        # Determine column layout based on file era
        # Pre-2025: prev_year, curr_prev_week, curr_week  (3 cols)
        # 2025+:    prev_year, curr_prev_week, curr_week, five_year_avg (4 cols)
        numeric_vals = [_coerce_pct(v) for v in data_values[:5]]

        # Find the actual data columns by matching against date_columns count
        n_dates = len([d for d in date_columns if d is not None])

        if has_five_year_avg and len(numeric_vals) >= 4:
            prev_year_pct = numeric_vals[0]
            prev_week_pct = numeric_vals[1]
            progress_pct = numeric_vals[2]
            five_year_avg_pct = numeric_vals[3]
        elif len(numeric_vals) >= 3:
            prev_year_pct = numeric_vals[0]
            prev_week_pct = numeric_vals[1]
            progress_pct = numeric_vals[2]
            five_year_avg_pct = None
        else:
            logger.debug(
                "    Insufficient data columns for row %r: %s",
                state_name, numeric_vals,
            )
            continue

        # Determine report_date from date_columns
        # The last date in date_columns is typically the current week end date
        actual_dates_list = [d for d in date_columns if d is not None]
        if actual_dates_list:
            report_date = actual_dates_list[-1]
            # week_start / week_end from the filename or date range
            if len(actual_dates_list) >= 2:
                week_end = actual_dates_list[-1]
                # The second-to-last date that's close to report_date is the
                # previous week, but the week_start for this week can be
                # derived (report_date - 6 days as approximation)
                from datetime import timedelta
                week_start = week_end - timedelta(days=6)
            else:
                week_end = actual_dates_list[0]
                from datetime import timedelta
                week_start = week_end - timedelta(days=6)
        else:
            continue

        record = {
            "report_date": report_date,
            "week_start": week_start,
            "week_end": week_end,
            "crop": current_crop_en,
            "crop_season": current_season,
            "activity": current_activity_en,
            "state": state_code_to_name(state_code) if state_code else "Total",
            "state_code": state_code or "",
            "is_total": is_total,
            "progress_pct": progress_pct,
            "prev_week_pct": prev_week_pct,
            "prev_year_pct": prev_year_pct,
            "five_year_avg_pct": five_year_avg_pct,
            "source_file": filepath.name,
        }
        records.append(record)

    wb.close()
    logger.info("  Parsed %d records from %s", len(records), filepath.name)
    return records


def parse_directory(directory: Path) -> list[dict]:
    """Parse all .xlsx files in a directory (recursively)."""
    all_records: list[dict] = []
    xlsx_files = sorted(directory.rglob("*.xlsx"))
    logger.info("Found %d xlsx files in %s", len(xlsx_files), directory)

    for fpath in xlsx_files:
        # Skip temporary/hidden files
        if fpath.name.startswith("~") or fpath.name.startswith("."):
            continue
        records = parse_file(fpath)
        all_records.extend(records)

    logger.info("Total records parsed: %d", len(all_records))
    return all_records
